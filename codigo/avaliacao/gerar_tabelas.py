import json
from pathlib import Path

from openpyxl import Workbook # cria um arquivo excel
from openpyxl.styles import ( # ferramentas do excel
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)
from openpyxl.utils import get_column_letter # converte numero de coluna em letra
from openpyxl.worksheet.table import ( #transformar os dados em tabela de excel
    Table,
    TableStyleInfo,
)



# CONFIGURAÇÕES


PASTAS = {
    "few_shot": {
        "enem": Path(
            "resultados/few_shot/enem"
        ),
        "gsm8k": Path(
            "resultados/few_shot/gsm8k"
        ),
    },

    "chain_of_thought": {
        "enem": Path(
            "resultados/chain_of_thought/enem"
        ),
        "gsm8k": Path(
            "resultados/chain_of_thought/gsm8k"
        ),
    },

    "self_consistency": {
        "enem": Path(
            "resultados/self_consistency/enem"
        ),
        "gsm8k": Path(
            "resultados/self_consistency/gsm8k"
        ),
    },
}


NOMES_METODOS = {
    "few_shot": "Few-Shot",
    "chain_of_thought": "Chain-of-Thought",
    "self_consistency": "Self-Consistency",
}



# ARQUIVOS DE SAÍDA


ARQ_GERAL = Path(
    "resultados/resultados_experimentos.xlsx"
)

ARQ_FEW_SHOT = Path(
    "resultados/few_shot/"
    "resultados_few_shot.xlsx"
)

ARQ_COT = Path(
    "resultados/chain_of_thought/"
    "resultados_chain_of_thought.xlsx"
)

ARQ_SC = Path(
    "resultados/self_consistency/"
    "resultados_self_consistency.xlsx"
)



# ESTILO


COR_CABECALHO = "1F4E78"
COR_TEXTO_CABECALHO = "FFFFFF"

COR_ACERTO = "C6EFCE"
COR_ERRO = "FFC7CE"
COR_EMPATE = "FFF2CC"

BORDA = Side(
    style="thin",
    color="D9E2F3"
)



# CARREGAR JSON


def carregar_json(caminho):

    try:

        with open(
            caminho,
            "r",
            encoding="utf-8"
        ) as arquivo:

            return json.load(arquivo)

    except Exception as erro:

        print(
            f"Erro ao carregar "
            f"{caminho}: {erro}"
        )

        return None



# CARREGAR TODAS AS EXECUÇÕES


def carregar_execucoes():

    execucoes = []

    for metodo, datasets in PASTAS.items():

        for dataset, pasta in datasets.items():

            if not pasta.exists():
                continue

            arquivos = sorted(
                pasta.glob("*.json")
            )

            for caminho in arquivos:

                dados = carregar_json(
                    caminho
                )

                if dados is None:
                    continue

                execucoes.append(
                    {
                        "metodo": metodo,
                        "dataset": dataset,
                        "arquivo": caminho.name,
                        "dados": dados,
                    }
                )

    return execucoes



# MÉTRICAS


def calcular_metricas(resultados):

    total = len(resultados)

    acertos = sum(
        1
        for resultado in resultados
        if resultado.get(
            "acertou",
            False
        )
    )

    erros = total - acertos

    precisao = (
        acertos / total
        if total > 0
        else 0
    )

    return (
        total,
        acertos,
        erros,
        precisao
    )



# LINHAS DO EXCEL GERAL


def criar_linhas_geral(execucoes):

    linhas = []

    for execucao in execucoes:

        dados = execucao["dados"]

        resultados = dados.get(
            "resultados",
            []
        )

        (
            total,
            acertos,
            erros,
            precisao
        ) = calcular_metricas(
            resultados
        )

        parametros = dados.get(
            "parametros",
            {}
        )

        linhas.append(
            [
                dados.get(
                    "id_execucao",
                    execucao["arquivo"]
                ),

                dados.get(
                    "data_execucao",
                    ""
                ),

                NOMES_METODOS.get(
                    execucao["metodo"],
                    execucao["metodo"]
                ),

                execucao[
                    "dataset"
                ].upper(),

                dados.get(
                    "modelo",
                    ""
                ),

                total,

                acertos,

                erros,

                precisao,

                parametros.get(
                    "limite",
                    total
                ),
            ]
        )

    linhas.sort(
        key=lambda linha: (
            linha[3],
            linha[5],
            linha[1],
            linha[2],
        )
    )

    return linhas



# LINHAS DETALHADAS DE UMA TÉCNICA


def criar_linhas_tecnica(
    execucoes,
    metodo
):

    linhas = []

    for execucao in execucoes:

        if execucao["metodo"] != metodo:
            continue

        dados = execucao["dados"]

        resultados = dados.get(
            "resultados",
            []
        )

        parametros = dados.get(
            "parametros",
            {}
        )

        id_execucao = dados.get(
            "id_execucao",
            execucao["arquivo"]
        )

        data_execucao = dados.get(
            "data_execucao",
            ""
        )

        modelo = dados.get(
            "modelo",
            ""
        )

        dataset = execucao[
            "dataset"
        ].upper()

        limite = parametros.get(
            "limite",
            len(resultados)
        )

        for resultado in resultados:

            acertou = resultado.get(
                "acertou",
                False
            )

            if dataset == "ENEM":

                identificador = (
                    f"{resultado.get('ano', '')}-"
                    f"{resultado.get('numero', '')}"
                )

                ano = resultado.get(
                    "ano",
                    ""
                )

                numero = resultado.get(
                    "numero",
                    ""
                )

                split = ""

                tem_imagem = (
                    "SIM"
                    if resultado.get(
                        "tem_imagem",
                        False
                    )
                    else "NÃO"
                )

            else:

                identificador = (
                    resultado.get(
                        "id",
                        ""
                    )
                )

                ano = ""
                numero = ""

                split = resultado.get(
                    "split",
                    ""
                )

                tem_imagem = ""

            linhas.append(
                [
                    id_execucao,

                    data_execucao,

                    dataset,

                    modelo,

                    limite,

                    identificador,

                    ano,

                    numero,

                    split,

                    resultado.get(
                        "gabarito",
                        ""
                    ),

                    resultado.get(
                        "resposta_modelo",
                        ""
                    ),

                    (
                        "ACERTO"
                        if acertou
                        else "ERRO"
                    ),

                    tem_imagem,

                    resultado.get(
                        "resposta_completa",
                        ""
                    ),

                    execucao["arquivo"],
                ]
            )

    return linhas



# LINHAS SELF-CONSISTENCY


def criar_linhas_self_consistency(
    execucoes
):

    linhas = []

    for execucao in execucoes:

        if (
            execucao["metodo"]
            != "self_consistency"
        ):
            continue

        dados = execucao["dados"]

        resultados = dados.get(
            "resultados",
            []
        )

        parametros = dados.get(
            "parametros",
            {}
        )

        id_execucao = dados.get(
            "id_execucao",
            execucao["arquivo"]
        )

        data_execucao = dados.get(
            "data_execucao",
            ""
        )

        modelo = dados.get(
            "modelo",
            ""
        )

        dataset = execucao[
            "dataset"
        ].upper()

        limite = parametros.get(
            "limite",
            len(resultados)
        )

        for resultado in resultados:

            acertou = resultado.get(
                "acertou",
                False
            )

            respostas = resultado.get(
                "respostas_extraidas",
                []
            )

            distribuicao = resultado.get(
                "distribuicao_votos",
                {}
            )

            if dataset == "ENEM":

                identificador = (
                    f"{resultado.get('ano', '')}-"
                    f"{resultado.get('numero', '')}"
                )

                ano = resultado.get(
                    "ano",
                    ""
                )

                numero = resultado.get(
                    "numero",
                    ""
                )

                split = ""

                tem_imagem = (
                    "SIM"
                    if resultado.get(
                        "tem_imagem",
                        False
                    )
                    else "NÃO"
                )

            else:

                identificador = (
                    resultado.get(
                        "id",
                        ""
                    )
                )

                ano = ""
                numero = ""

                split = resultado.get(
                    "split",
                    ""
                )

                tem_imagem = ""

            linhas.append(
                [
                    id_execucao,

                    data_execucao,

                    dataset,

                    modelo,

                    limite,

                    identificador,

                    ano,

                    numero,

                    split,

                    resultado.get(
                        "gabarito",
                        ""
                    ),

                    resultado.get(
                        "resposta_modelo",
                        ""
                    ),

                    (
                        "ACERTO"
                        if acertou
                        else "ERRO"
                    ),

                    tem_imagem,

                    respostas[0]
                    if len(respostas) > 0
                    else "",

                    respostas[1]
                    if len(respostas) > 1
                    else "",

                    respostas[2]
                    if len(respostas) > 2
                    else "",

                    respostas[3]
                    if len(respostas) > 3
                    else "",

                    respostas[4]
                    if len(respostas) > 4
                    else "",

                    (
                        json.dumps(
                            distribuicao,
                            ensure_ascii=False
                        )
                        if distribuicao
                        else ""
                    ),

                    resultado.get(
                        "votos_resposta_modelo",
                        ""
                    ),

                    (
                        "SIM"
                        if resultado.get(
                            "houve_empate",
                            False
                        )
                        else "NÃO"
                    ),

                    execucao["arquivo"],
                ]
            )

    return linhas



# FORMATAÇÃO


def formatar_aba(
    ws,
    coluna_resultado=None,
    coluna_precisao=None,
    coluna_empate=None
):

    # Cabeçalho
    for celula in ws[1]:

        celula.font = Font(
            bold=True,
            color=COR_TEXTO_CABECALHO
        )

        celula.fill = PatternFill(
            fill_type="solid",
            fgColor=COR_CABECALHO
        )

        celula.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.freeze_panes = "A2"

    ws.auto_filter.ref = (
        ws.dimensions
    )

    # Dados
    for row in ws.iter_rows(
        min_row=2
    ):

        for celula in row:

            celula.alignment = Alignment(
                vertical="top",
                wrap_text=True
            )

            celula.border = Border(
                bottom=BORDA
            )

    # Precisão
    if coluna_precisao:

        for linha in range(
            2,
            ws.max_row + 1
        ):

            ws[
                f"{coluna_precisao}{linha}"
            ].number_format = "0.00%"

    # Acerto / erro
    if coluna_resultado:

        for linha in range(
            2,
            ws.max_row + 1
        ):

            celula = ws[
                f"{coluna_resultado}{linha}"
            ]

            if celula.value == "ACERTO":

                celula.fill = PatternFill(
                    fill_type="solid",
                    fgColor=COR_ACERTO
                )

            elif celula.value == "ERRO":

                celula.fill = PatternFill(
                    fill_type="solid",
                    fgColor=COR_ERRO
                )

    # Empate
    if coluna_empate:

        for linha in range(
            2,
            ws.max_row + 1
        ):

            celula = ws[
                f"{coluna_empate}{linha}"
            ]

            if celula.value == "SIM":

                celula.fill = PatternFill(
                    fill_type="solid",
                    fgColor=COR_EMPATE
                )

    # Largura das colunas
    for coluna in range(
        1,
        ws.max_column + 1
    ):

        letra = get_column_letter(
            coluna
        )

        titulo = ws.cell(
            row=1,
            column=coluna
        ).value

        if titulo == "Resposta completa":
            largura = 55

        elif titulo == "ID execução":
            largura = 42

        elif titulo == "Arquivo origem":
            largura = 48

        elif titulo == "Distribuição votos":
            largura = 28

        else:
            largura = 18

        ws.column_dimensions[
            letra
        ].width = largura



# ADICIONAR TABELA DO EXCEL


def adicionar_tabela(
    ws,
    nome
):

    if ws.max_row < 2:
        return

    ultima_coluna = (
        get_column_letter(
            ws.max_column
        )
    )

    referencia = (
        f"A1:"
        f"{ultima_coluna}"
        f"{ws.max_row}"
    )

    tabela = Table(
        displayName=nome,
        ref=referencia
    )

    estilo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )

    tabela.tableStyleInfo = estilo

    ws.add_table(
        tabela
    )



# CRIAR EXCEL GERAL


def criar_excel_geral(
    execucoes
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Resultados"

    cabecalhos = [
        "ID execução",
        "Data",
        "Método",
        "Dataset",
        "Modelo",
        "Total",
        "Acertos",
        "Erros",
        "Precisão",
        "Limite",
    ]

    ws.append(
        cabecalhos
    )

    for linha in criar_linhas_geral(
        execucoes
    ):
        ws.append(linha)

    formatar_aba(
        ws,
        coluna_precisao="I"
    )

    adicionar_tabela(
        ws,
        "TabelaResultadosGerais"
    )

    ARQ_GERAL.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(
        ARQ_GERAL
    )

    print(
        f"Criado: {ARQ_GERAL}"
    )



# CRIAR EXCEL FEW-SHOT


def criar_excel_few_shot(
    execucoes
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Resultados"

    cabecalhos = [
        "ID execução",
        "Data",
        "Dataset",
        "Modelo",
        "Limite",
        "ID questão",
        "Ano",
        "Questão",
        "Split",
        "Gabarito",
        "Resposta modelo",
        "Resultado",
        "Tem imagem",
        "Resposta completa",
        "Arquivo origem",
    ]

    ws.append(
        cabecalhos
    )

    linhas = criar_linhas_tecnica(
        execucoes,
        "few_shot"
    )

    for linha in linhas:
        ws.append(linha)

    formatar_aba(
        ws,
        coluna_resultado="L"
    )

    adicionar_tabela(
        ws,
        "TabelaFewShot"
    )

    ARQ_FEW_SHOT.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(
        ARQ_FEW_SHOT
    )

    print(
        f"Criado: {ARQ_FEW_SHOT}"
    )



# CRIAR EXCEL COT


def criar_excel_cot(
    execucoes
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Resultados"

    cabecalhos = [
        "ID execução",
        "Data",
        "Dataset",
        "Modelo",
        "Limite",
        "ID questão",
        "Ano",
        "Questão",
        "Split",
        "Gabarito",
        "Resposta modelo",
        "Resultado",
        "Tem imagem",
        "Raciocínio completo",
        "Arquivo origem",
    ]

    ws.append(
        cabecalhos
    )

    linhas = criar_linhas_tecnica(
        execucoes,
        "chain_of_thought"
    )

    for linha in linhas:
        ws.append(linha)

    formatar_aba(
        ws,
        coluna_resultado="L"
    )

    # Raciocínio completo está na coluna N
    ws.column_dimensions[
        "N"
    ].width = 60

    adicionar_tabela(
        ws,
        "TabelaChainOfThought"
    )

    ARQ_COT.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(
        ARQ_COT
    )

    print(
        f"Criado: {ARQ_COT}"
    )



# CRIAR EXCEL SELF-CONSISTENCY


def criar_excel_sc(
    execucoes
):

    wb = Workbook()

    ws = wb.active

    ws.title = "Resultados"

    cabecalhos = [
        "ID execução",
        "Data",
        "Dataset",
        "Modelo",
        "Limite",
        "ID questão",
        "Ano",
        "Questão",
        "Split",
        "Gabarito",
        "Resposta final",
        "Resultado",
        "Tem imagem",
        "Amostra 1",
        "Amostra 2",
        "Amostra 3",
        "Amostra 4",
        "Amostra 5",
        "Distribuição votos",
        "Votos vencedor",
        "Houve empate",
        "Arquivo origem",
    ]

    ws.append(
        cabecalhos
    )

    linhas = (
        criar_linhas_self_consistency(
            execucoes
        )
    )

    for linha in linhas:
        ws.append(linha)

    formatar_aba(
        ws,
        coluna_resultado="L",
        coluna_empate="U"
    )

    adicionar_tabela(
        ws,
        "TabelaSelfConsistency"
    )

    ARQ_SC.parent.mkdir(
        parents=True,
        exist_ok=True
    )

    wb.save(
        ARQ_SC
    )

    print(
        f"Criado: {ARQ_SC}"
    )



# MAIN


def main():

    print()
    print("=" * 70)
    print(
        "GERANDO PLANILHAS DOS EXPERIMENTOS"
    )
    print("=" * 70)
    print()

    execucoes = carregar_execucoes()

    print(
        f"Execuções encontradas: "
        f"{len(execucoes)}"
    )

    if not execucoes:

        print(
            "Nenhuma execução encontrada."
        )

        return

    print()

    criar_excel_geral(
        execucoes
    )

    criar_excel_few_shot(
        execucoes
    )

    criar_excel_cot(
        execucoes
    )

    criar_excel_sc(
        execucoes
    )

    print()
    print("=" * 70)
    print(
        "PLANILHAS GERADAS COM SUCESSO"
    )
    print("=" * 70)

    print()
    print(
        "1. resultados/"
        "resultados_experimentos.xlsx"
    )

    print(
        "2. resultados/few_shot/"
        "resultados_few_shot.xlsx"
    )

    print(
        "3. resultados/chain_of_thought/"
        "resultados_chain_of_thought.xlsx"
    )

    print(
        "4. resultados/self_consistency/"
        "resultados_self_consistency.xlsx"
    )

    print()



# EXECUÇÃO


if __name__ == "__main__":
    main()